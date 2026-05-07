import math
import random
import time
import tkinter as tk
from tkinter import filedialog

from openpyxl import load_workbook
import pyglet
from pyglet import gl
from pyglet.window import key, mouse


WIDTH = 1100
HEIGHT = 720
PANEL_W = 330

BG = (14, 17, 21)
PANEL = (31, 35, 40)
PANEL_LINE = (71, 78, 86)
TEXT = (235, 239, 244, 255)
MUTED = (157, 166, 178, 255)
BUTTON = (52, 61, 69)
BUTTON_HOVER = (67, 78, 88)
START = (31, 140, 108)
START_HOVER = (38, 168, 130)
RESET = (184, 56, 61)
RESET_HOVER = (219, 69, 74)
INPUT_BG = (18, 21, 25)
LIST_BG = (19, 22, 26)
SELECTED = (46, 87, 92)

WHEEL_COLORS = [
    (240, 87, 87),
    (250, 174, 61),
    (89, 189, 125),
    (51, 163, 219),
    (138, 112, 212),
    (227, 110, 184),
    (120, 199, 192),
    (219, 148, 89),
]


class Rect:
    def __init__(self, x, y, w, h):
        self.x = x
        self.y = y
        self.w = w
        self.h = h

    @property
    def cx(self):
        return self.x + self.w / 2

    @property
    def cy(self):
        return self.y + self.h / 2

    def contains(self, x, y):
        return self.x <= x <= self.x + self.w and self.y <= y <= self.y + self.h


class Button:
    def __init__(self, rect, label, kind="normal"):
        self.rect = rect
        self.label = label
        self.kind = kind


class SpinnerWindow(pyglet.window.Window):
    def __init__(self):
        super().__init__(WIDTH, HEIGHT, "OpenGL Text Spinner", resizable=True, vsync=True)
        self.set_minimum_size(760, 520)
        gl.glClearColor(BG[0] / 255, BG[1] / 255, BG[2] / 255, 1)
        gl.glEnable(gl.GL_BLEND)
        gl.glBlendFunc(gl.GL_SRC_ALPHA, gl.GL_ONE_MINUS_SRC_ALPHA)

        self.items = ["Alice", "Bob", "Charlie", "Diana", "Winner"]
        self.input_text = ""
        self.selected_index = None
        self.message = "Add names, then press Start."
        self.winner = None

        self.rotation = 0.0
        self.spin_target = 0.0
        self.start_rotation = 0.0
        self.spin_start = 0.0
        self.spin_duration = 0.0
        self.spinning = False
        self.hover = (0, 0)

        self.layout()
        pyglet.clock.schedule_interval(self.update, 1 / 60)

    def layout(self):
        left = self.width - PANEL_W + 24
        self.input_rect = Rect(left, self.height - 158, PANEL_W - 48, 42)
        self.add_button = Button(Rect(left, self.height - 212, 134, 42), "Add")
        self.remove_button = Button(Rect(left + 146, self.height - 212, 134, 42), "Remove")
        self.upload_box = Rect(left, self.height - 282, PANEL_W - 48, 58)
        self.upload_button = Button(Rect(left + 142, self.height - 270, 138, 34), "Upload Excel")
        self.list_rect = Rect(left, 170, PANEL_W - 48, max(90, self.height - 500))
        self.start_button = Button(Rect(left, 90, 134, 48), "Start", "start")
        self.reset_button = Button(Rect(left + 146, 90, 134, 48), "Reset", "reset")
        self.clear_button = Button(Rect(left, 30, PANEL_W - 48, 42), "Clear List")

    def on_resize(self, width, height):
        super().on_resize(width, height)
        self.layout()

    def on_mouse_motion(self, x, y, dx, dy):
        self.hover = (x, y)

    def on_mouse_press(self, x, y, button, modifiers):
        if button != mouse.LEFT:
            return
        if self.add_button.rect.contains(x, y):
            self.add_item()
        elif self.remove_button.rect.contains(x, y):
            self.remove_selected()
        elif self.upload_button.rect.contains(x, y):
            self.upload_excel()
        elif self.start_button.rect.contains(x, y):
            self.start_spin()
        elif self.reset_button.rect.contains(x, y):
            self.reset_spin()
        elif self.clear_button.rect.contains(x, y):
            self.clear_items()
        elif self.list_rect.contains(x, y):
            row = int((self.list_rect.y + self.list_rect.h - y) // 34)
            if 0 <= row < len(self.visible_items()):
                self.selected_index = row

    def on_text(self, text):
        if text.isprintable() and len(self.input_text) < 40:
            self.input_text += text

    def on_key_press(self, symbol, modifiers):
        if symbol == key.ESCAPE:
            self.close()
        elif symbol == key.ENTER:
            self.add_item()
        elif symbol == key.BACKSPACE:
            self.input_text = self.input_text[:-1]
        elif symbol == key.DELETE:
            self.remove_selected()

    def add_item(self):
        value = self.input_text.strip()
        if not value:
            return
        self.items.append(value)
        self.input_text = ""
        self.selected_index = len(self.items) - 1
        self.winner = None
        self.message = f"Added {value}."

    def remove_selected(self):
        if self.selected_index is None or not (0 <= self.selected_index < len(self.items)):
            return
        removed = self.items.pop(self.selected_index)
        self.selected_index = None
        self.winner = None
        self.message = f"Removed {removed}."

    def clear_items(self):
        self.items.clear()
        self.selected_index = None
        self.input_text = ""
        self.winner = None
        self.message = "List cleared."

    def upload_excel(self):
        try:
            root = tk.Tk()
            root.withdraw()
            root.attributes("-topmost", True)
            path = filedialog.askopenfilename(
                title="Select Excel file",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            )
            root.destroy()
        except Exception as exc:
            self.message = f"Could not open file picker: {exc}"
            return

        if not path:
            self.message = "Upload cancelled."
            return

        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook.active
            values = []
            for row in sheet.iter_rows(min_col=1, max_col=1, values_only=True):
                value = row[0]
                if value is None:
                    continue
                text = str(value).strip()
                if text:
                    values.append(text)
            workbook.close()
        except Exception as exc:
            self.message = f"Excel upload failed: {exc}"
            return

        if not values:
            self.message = "No values found in the first column."
            return

        self.items = values
        self.selected_index = None
        self.winner = None
        self.rotation = 0.0
        self.message = f"Loaded {len(values)} values from Excel."

    def reset_spin(self):
        self.spinning = False
        self.rotation = 0.0
        self.winner = None
        self.message = "Ready."

    def start_spin(self):
        if self.spinning:
            return
        if len(self.items) < 2:
            self.message = "Add at least two values before spinning."
            return

        winner_index = random.randrange(len(self.items))
        segment = 360.0 / len(self.items)
        segment_center = winner_index * segment + segment / 2
        current = self.rotation % 360.0
        desired_mod = (-segment_center) % 360.0
        delta = (desired_mod - current) % 360.0

        self.start_rotation = self.rotation
        self.spin_target = self.rotation + random.randint(5, 8) * 360.0 + delta
        self.spin_duration = random.uniform(3.2, 4.8)
        self.spin_start = time.perf_counter()
        self.spinning = True
        self.winner = None
        self.message = "Spinning..."

    def update(self, dt):
        if not self.spinning:
            return
        t = min(1.0, (time.perf_counter() - self.spin_start) / self.spin_duration)
        eased = 1 - (1 - t) ** 4
        self.rotation = self.start_rotation + (self.spin_target - self.start_rotation) * eased
        if t >= 1:
            self.spinning = False
            self.rotation %= 360.0
            self.selected_index = self.current_winner_index()
            self.winner = self.items[self.selected_index]
            self.message = f"Selected: {self.winner}"

    def current_winner_index(self):
        segment = 360.0 / len(self.items)
        wheel_angle_at_pointer = (-self.rotation) % 360.0
        return int(wheel_angle_at_pointer // segment) % len(self.items)

    def visible_items(self):
        max_rows = int(self.list_rect.h // 34)
        return self.items[:max_rows]

    def draw_rect(self, rect, color):
        pyglet.shapes.Rectangle(rect.x, rect.y, rect.w, rect.h, color=color).draw()

    def draw_label(self, text, x, y, size=16, color=TEXT, bold=False, anchor_x="left", anchor_y="baseline"):
        pyglet.text.Label(
            text,
            font_name="Segoe UI",
            font_size=size,
            color=color,
            x=x,
            y=y,
            anchor_x=anchor_x,
            anchor_y=anchor_y,
        ).draw()

    def draw_button(self, button):
        hovered = button.rect.contains(*self.hover)
        if button.kind == "start":
            color = START_HOVER if hovered else START
        elif button.kind == "reset":
            color = RESET_HOVER if hovered else RESET
        else:
            color = BUTTON_HOVER if hovered else BUTTON
        self.draw_rect(button.rect, color)
        self.draw_label(button.label, button.rect.cx, button.rect.cy, 16, TEXT, True, "center", "center")

    def draw_wheel(self):
        center_x = (self.width - PANEL_W) / 2
        center_y = self.height / 2 + 18
        radius = min(self.height * 0.37, (self.width - PANEL_W) * 0.39)

        self.draw_label("Random Picker", center_x, self.height - 42, 30, TEXT, True, "center", "center")
        self.draw_label(self.message, center_x, self.height - 78, 16, MUTED, False, "center", "center")

        count = max(1, len(self.items))
        segment = 360.0 / count
        for i in range(count):
            start = math.radians(i * segment + self.rotation)
            end = math.radians((i + 1) * segment + self.rotation)
            color = WHEEL_COLORS[i % len(WHEEL_COLORS)] if self.items else (72, 78, 88)
            steps = max(10, int(segment / 3))
            previous = (
                center_x + math.cos(start) * radius,
                center_y + math.sin(start) * radius,
            )
            for step in range(1, steps + 1):
                angle = start + (end - start) * step / steps
                current = (
                    center_x + math.cos(angle) * radius,
                    center_y + math.sin(angle) * radius,
                )
                pyglet.shapes.Triangle(
                    center_x,
                    center_y,
                    previous[0],
                    previous[1],
                    current[0],
                    current[1],
                    color=color,
                ).draw()
                previous = current

        for i, value in enumerate(self.items or ["Add values"]):
            angle = math.radians(i * segment + segment / 2 + self.rotation)
            label = value if len(value) <= 18 else value[:15] + "..."
            x = center_x + math.cos(angle) * radius * 0.55
            y = center_y + math.sin(angle) * radius * 0.55
            self.draw_label(label, x, y, 16, (20, 24, 28, 255), True, "center", "center")

        pointer = [
            (center_x + radius + 18, center_y),
            (center_x + radius + 58, center_y + 18),
            (center_x + radius + 58, center_y - 18),
        ]
        pyglet.shapes.Triangle(*pointer[0], *pointer[1], *pointer[2], color=(255, 255, 255)).draw()
        pyglet.shapes.Circle(center_x, center_y, 43, color=(15, 18, 21)).draw()

        if self.winner:
            self.draw_label("Winner", center_x, center_y + 12, 12, MUTED, False, "center", "center")
            display = self.winner if len(self.winner) <= 20 else self.winner[:17] + "..."
            self.draw_label(display, center_x, center_y - 13, 16, TEXT, True, "center", "center")
        else:
            self.draw_label("Ready", center_x, center_y, 16, TEXT, True, "center", "center")

    def draw_panel(self):
        self.draw_rect(Rect(self.width - PANEL_W, 0, PANEL_W, self.height), PANEL)
        self.draw_rect(Rect(self.width - PANEL_W, 0, 2, self.height), PANEL_LINE)

        left = self.width - PANEL_W + 24
        self.draw_label("Text Spinner", left, self.height - 44, 27, TEXT, True)
        self.draw_label("Input", left, self.height - 100, 15, TEXT, True)
        self.draw_rect(self.input_rect, INPUT_BG)
        value = self.input_text if self.input_text else "Type a name or text value"
        color = TEXT if self.input_text else MUTED
        self.draw_label(value, self.input_rect.x + 12, self.input_rect.y + 13, 15, color)

        self.draw_button(self.add_button)
        self.draw_button(self.remove_button)

        self.draw_rect(self.upload_box, (24, 29, 34))
        self.draw_label("file with only a column", self.upload_box.x + 10, self.upload_box.y + 21, 13, MUTED)
        self.draw_button(self.upload_button)

        self.draw_label(f"Values ({len(self.items)})", left, self.list_rect.y + self.list_rect.h + 18, 15, TEXT, True)
        self.draw_rect(self.list_rect, LIST_BG)
        for i, item in enumerate(self.visible_items()):
            y = self.list_rect.y + self.list_rect.h - 34 * (i + 1)
            if i == self.selected_index:
                self.draw_rect(Rect(self.list_rect.x, y, self.list_rect.w, 34), SELECTED)
            label = item if len(item) <= 24 else item[:21] + "..."
            self.draw_label(label, self.list_rect.x + 10, y + 10, 13, TEXT)

        self.draw_button(self.start_button)
        self.draw_button(self.reset_button)
        self.draw_button(self.clear_button)

    def on_draw(self):
        self.clear()
        self.draw_wheel()
        self.draw_panel()


if __name__ == "__main__":
    SpinnerWindow()
    pyglet.app.run()
